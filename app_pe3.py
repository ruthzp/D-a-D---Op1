import os
import shutil
import tempfile
import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
from copy import copy, deepcopy

from funciones_asistencia import generar_asistencia
from funciones_op1 import generar_op1
from funciones_personal import generar_personal
from funciones_cajas_sede import generar_cajas_sede


# ========================
# CONFIGURACIÓN
# ========================
st.set_page_config(page_title="Sistema PE", layout="wide")

for key in [
    "asistencia_generada",
    "op1_generada",
    "personal_generada",
    "cajas_sede_generada",
]:
    if key not in st.session_state:
        st.session_state[key] = None


# ========================
# PLANTILLA BASE
# ========================
def _get_plantilla_path():
    base_dir = os.path.dirname(os.path.abspath(__file__))
    candidatos = [
        os.path.join(base_dir, "plantillas", "Op1 - Reporte.xlsx"),
        os.path.join(base_dir, "Op1 - Reporte.xlsx"),
    ]
    for ruta in candidatos:
        if os.path.exists(ruta):
            return ruta
    raise FileNotFoundError("❌ No se encontró la plantilla ‘Op1 - Reporte.xlsx’.")
    

PLANTILLA_PATH = _get_plantilla_path()


def get_temp_copy():
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    shutil.copyfile(PLANTILLA_PATH, tmp.name)
    return tmp.name


# ========================
# CLASIFICADOR (CORREGIDO)
# ========================
def clasificar_archivos(lista):
    res = {
        "asc": None,
        "nom": None,
        "asc_mindef": None,
        "asc_inst": None,
        "nom_inst": None,
        "asc_fa": None,
        "mindef_inst": None,
        "asc_personal": None,
        "asc_cajas_sede": None,
    }

    for f in lista:
        nombre = f.name.upper().replace(" ", "")

        # 1) PERSONAL
        if "PERSONAL" in nombre and "ASC" in nombre:
            res["asc_personal"] = f
            continue

        # 2) CAJAS - SEDE
        if "CAJAS" in nombre and "SEDE" in nombre and "ASC" in nombre:
            res["asc_cajas_sede"] = f
            continue

        # 3) MINDEF – POSTULANTES
        if "MINDEF" in nombre and "POSTULANTE" in nombre:
            res["asc_mindef"] = f
            continue

        # 4) POSTULANTES (ASC / NOM)
        if "POSTULANTE" in nombre:
            if "ASC" in nombre:
                res["asc"] = f
            elif "NOM" in nombre:
                res["nom"] = f
            continue

        # 5) MINDEF – INSTRUMENTOS
        if "MINDEF" in nombre and ("INSTRUMENTO" in nombre or "INSTRUMENTOS" in nombre):
            res["mindef_inst"] = f
            continue

        # 6) INSTRUMENTOS (ASC / NOM)
        if "INSTRUMENTO" in nombre or "INSTRUMENTOS" in nombre:
            if "ASC" in nombre:
                res["asc_inst"] = f
            elif "NOM" in nombre:
                res["nom_inst"] = f
            continue

        # 7) ASC – FA
        if "FA" in nombre and "ASC" in nombre:
            res["asc_fa"] = f
            continue

    return res


# ========================
# COPIAR HOJA COMPLETA
# ========================
def copiar_hoja_completa(ws_src, ws_dest):

    for row in ws_src.iter_rows():
        for cell in row:
            new_cell = ws_dest.cell(row=cell.row, column=cell.col_idx, value=cell.value)

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.alignment = copy(cell.alignment)

    for merged in ws_src.merged_cells.ranges:
        ws_dest.merge_cells(str(merged))

    for col, dim in ws_src.column_dimensions.items():
        ws_dest.column_dimensions[col].width = dim.width

    for r, dim in ws_src.row_dimensions.items():
        ws_dest.row_dimensions[r].height = dim.height

    try:
        if getattr(ws_src, "conditional_formatting", None):
            ws_dest.conditional_formatting = deepcopy(ws_src.conditional_formatting)
    except Exception:
        pass


# ========================
# COMBINAR REPORTES
# ========================
def combinar_reportes(plantilla, asistencia=None, op1=None,
                      personal=None, cajas_sede=None):

    wb_final = load_workbook(plantilla)

    reportes = {
        "ASISTENCIA": asistencia,
        "OP1": op1,
        "PERSONAL": personal,
        "CAJAS-SEDE": cajas_sede
    }

    hojas_protegidas = ["DIC"]
    for hoja in wb_final.sheetnames[:]:
        if hoja not in hojas_protegidas:
            del wb_final[hoja]

    for nombre, archivo_bytes in reportes.items():
        if not archivo_bytes:
            continue

        wb_src = load_workbook(archivo_bytes)
        ws_src = wb_src.active

        ws_new = wb_final.create_sheet(nombre)
        copiar_hoja_completa(ws_src, ws_new)

    if "DIC" in wb_final.sheetnames:
        dic = wb_final["DIC"]
        wb_final._sheets.remove(dic)
        wb_final._sheets.insert(0, dic)

    try:
        wb_final.calculation_properties.fullCalcOnLoad = True
    except:
        pass

    out = BytesIO()
    wb_final.save(out)
    out.seek(0)
    return out


# ========================
# UI STREAMLIT
# ========================
st.title("📊 Sistema de Generación de Reportes PE")
st.caption("Sube los archivos Excel. Se clasificarán automáticamente.")

with st.expander("📁 Subir archivos", expanded=True):
    archivos = st.file_uploader(
        "Selecciona archivos (.xlsx)",
        type=["xlsx"],
        accept_multiple_files=True
    )

clasificados = clasificar_archivos(archivos) if archivos else {}

if archivos:
    with st.expander("📄 Archivos detectados"):
        cols = st.columns(3)
        for i, (k, v) in enumerate(clasificados.items()):
            if v:
                with cols[i % 3]:
                    st.markdown(
                        f"✅ **{k.upper()}**<br><small>{v.name}</small>",
                        unsafe_allow_html=True
                    )


# ========================
# GENERAR REPORTES
# ========================
if archivos:
    st.subheader("⚙️ Generar")
    c1, c2, c3, c4 = st.columns(4)

    # PERSONAL
    with c1:
        if st.button("👥 PERSONAL", disabled=clasificados.get("asc_personal") is None):
            tmp = get_temp_copy()
            generar_personal(tmp, clasificados["asc_personal"])
            st.toast("PERSONAL generado", icon="👥")

    # CAJAS-SEDE
    with c2:
        if st.button("🏢 CAJAS-SEDE", disabled=clasificados.get("asc_cajas_sede") is None):
            tmp = get_temp_copy()
            generar_cajas_sede(tmp, clasificados["asc_cajas_sede"])
            st.toast("CAJAS-SEDE generado", icon="🏢")

    # ASISTENCIA
    with c3:
        if st.button("🟢 ASISTENCIA", disabled=not (clasificados.get("asc") and clasificados.get("nom"))):
            tmp = get_temp_copy()
            generar_asistencia(
                tmp,
                clasificados["asc"],
                clasificados["nom"],
                clasificados.get("asc_mindef")
            )
            st.toast("ASISTENCIA generado", icon="🟢")

    # OP1
    with c4:
        if st.button("🟦 OP1",
                     disabled=not (clasificados.get("asc_inst")
                                   and clasificados.get("nom_inst")
                                   and clasificados.get("asc_fa"))):
            tmp = get_temp_copy()
            generar_op1(
                tmp,
                clasificados["asc_fa"],
                clasificados["asc_inst"],
                clasificados["nom_inst"],
                clasificados.get("mindef_inst")
            )
            st.toast("OP1 generado", icon="🟦")


# ========================
# DESCARGAS INDIVIDUALES
# ========================
st.divider()
st.subheader("⬇️ Descargas individuales")

cols = st.columns(4)

if st.session_state["personal_generada"]:
    cols[0].download_button(
        "PERSONAL",
        st.session_state["personal_generada"],
        file_name="PE - PERSONAL.xlsx"
    )

if st.session_state["cajas_sede_generada"]:
    cols[1].download_button(
        "CAJAS-SEDE",
        st.session_state["cajas_sede_generada"],
        file_name="PE - CAJAS_SEDE.xlsx"
    )

if st.session_state["asistencia_generada"]:
    cols[2].download_button(
        "ASISTENCIA",
        st.session_state["asistencia_generada"],
        file_name="PE - ASISTENCIA.xlsx"
    )

if st.session_state["op1_generada"]:
    cols[3].download_button(
        "OP1",
        st.session_state["op1_generada"],
        file_name="PE - OP1.xlsx"
    )


# ========================
# REPORTE FINAL
# ========================
st.divider()
st.subheader("📘 Reporte Final")

hay_reportes = (
    st.session_state["personal_generada"]
    or st.session_state["cajas_sede_generada"]
    or st.session_state["asistencia_generada"]
    or st.session_state["op1_generada"]
)

if hay_reportes:
    combinado = combinar_reportes(
        PLANTILLA_PATH,
        asistencia=st.session_state["asistencia_generada"],
        op1=st.session_state["op1_generada"],
        personal=st.session_state["personal_generada"],
        cajas_sede=st.session_state["cajas_sede_generada"]
    )

    st.download_button(
        "⬇️ Descargar Reporte Final",
        combinado,
        file_name="PE - Reporte_Final.xlsx"
    )
else:
    st.info("Genera al menos un reporte para combinarlo.", icon="ℹ️")