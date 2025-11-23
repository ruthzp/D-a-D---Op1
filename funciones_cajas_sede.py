import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import CellIsRule
from io import BytesIO


# ---------------------------------------------------------
# NORMALIZADOR
# ---------------------------------------------------------
def limpiar(s):
    if pd.isna(s):
        return ""
    return (
        str(s)
        .replace("\xa0", " ")
        .replace("\t", " ")
        .replace("\n", " ")
        .strip()
        .upper()
    )


def _to_int(v):
    """Convierte valores a enteros sin romper."""
    try:
        if pd.isna(v):
            return 0
    except Exception:
        pass
    try:
        return int(v)
    except Exception:
        try:
            return int(float(v))
        except Exception:
            return 0


# ---------------------------------------------------------
# DETECTAR ENCABEZADO PARA ASC – CAJAS – SEDE
# ---------------------------------------------------------
def detectar_fila_encabezados_cajas_sede(df_raw):

    for i in range(min(200, len(df_raw))):
        fila = df_raw.iloc[i].astype(str).apply(limpiar)
        valores = list(fila.values)

        if (
            "SEDE OPERATIVA" in valores
            and "TIPO" in valores
            and "TOTAL INVENTARIO IMPRENTA" in valores
            and "INGRESO" in valores
            and "SALIDA" in valores
        ):
            return i

    raise ValueError("No se encontró encabezado en ASC - CAJAS SEDE.")


# ---------------------------------------------------------
# CARGAR ASC – CAJAS – SEDE
# ---------------------------------------------------------
def cargar_asc_cajas_sede(archivo_asc):

    df_raw = pd.read_excel(archivo_asc, sheet_name="Reporte", header=None)
    fila = detectar_fila_encabezados_cajas_sede(df_raw)

    df = pd.read_excel(
        archivo_asc,
        sheet_name="Reporte",
        header=fila
    )

    df.columns = [limpiar(c) for c in df.columns]

    obligatorias = [
        "SEDE OPERATIVA",
        "TIPO",
        "TOTAL INVENTARIO IMPRENTA",
        "INGRESO",
        "SALIDA",
    ]

    faltan = [c for c in obligatorias if c not in df.columns]
    if faltan:
        raise ValueError(f"Faltan columnas obligatorias en ASC - CAJAS SEDE: {faltan}")

    df["SEDE OPERATIVA"] = df["SEDE OPERATIVA"].apply(limpiar)
    df["TIPO"] = df["TIPO"].apply(limpiar)

    return df


# ---------------------------------------------------------
# CLASIFICAR TIPO
# ---------------------------------------------------------
def clasificar_tipo(tipo):
    t = limpiar(tipo)

    if "APLIC" in t:
        return "INSTRUMENTO"

    if "ADIC" in t:
        return "ADICIONAL"

    if "CAND" in t:
        return "CANDADO"

    return None


# ---------------------------------------------------------
# GENERAR CAJAS-SEDE
# ---------------------------------------------------------
def generar_cajas_sede(ruta_plantilla_temp, archivo_asc_cajas_sede):

    try:
        with st.spinner("Generando hoja CAJAS-SEDE..."):

            # --- 1) Cargar ASC
            df = cargar_asc_cajas_sede(archivo_asc_cajas_sede)

            # --- 2) Agrupar por SEDE + TIPO
            index = {}

            for _, row in df.iterrows():
                tipo_cl = clasificar_tipo(row["TIPO"])
                if not tipo_cl:
                    continue

                sede = row["SEDE OPERATIVA"]

                key = (sede, tipo_cl)

                if key not in index:
                    index[key] = {"T": 0, "I": 0, "S": 0}

                index[key]["T"] += _to_int(row["TOTAL INVENTARIO IMPRENTA"])
                index[key]["I"] += _to_int(row["INGRESO"])
                index[key]["S"] += _to_int(row["SALIDA"])

            # --- 3) Cargar plantilla
            wb = load_workbook(ruta_plantilla_temp)
            ws = wb["CAJAS-SEDE"]

            # Mapear encabezados fila 1
            header_map = {}
            for c in ws[1]:
                if c.value:
                    header_map[limpiar(c.value)] = c.column_letter

            # Identificar columnas pero NO escribir en C, D, E
            col_sede = header_map.get("SEDE")

            # Columnas T (C, D, E) — NO SE MODIFICAN
            # col_T_INSTR = header_map["CAJA DE INSTRUMENTO DE APLICACIÓN[T]"]
            # col_T_ADIC = header_map["CAJA DE INSTRUMENTO ADICIONAL[T]"]
            # col_T_CAND = header_map["CAJA DE CANDADO[T]"]

            # Columnas I (estas SÍ se llenan)
            col_I_INSTR = header_map["CAJA DE INSTRUMENTO DE APLICACIÓN-I"]
            col_I_ADIC  = header_map["CAJA DE INSTRUMENTO ADICIONAL-I"]
            col_I_CAND  = header_map["CAJA DE CANDADO-I"]

            # Columnas I%
            col_IP_INSTR = header_map["CAJA DE INSTRUMENTO DE APLICACIÓN-I[P]"]
            col_IP_ADIC  = header_map["CAJA DE INSTRUMENTO ADICIONAL-I[P]"]
            col_IP_CAND  = header_map["CAJA DE CANDADO-I[P]"]

            # Columnas S (se llenan)
            col_S_INSTR = header_map["CAJA DE INSTRUMENTO DE APLICACIÓN-S"]
            col_S_ADIC  = header_map["CAJA DE INSTRUMENTO ADICIONAL-S"]
            col_S_CAND  = header_map["CAJA DE CANDADO-S"]

            # Columnas S%
            col_SP_INSTR = header_map["CAJA DE INSTRUMENTO DE APLICACIÓN-S[P]"]
            col_SP_ADIC  = header_map["CAJA DE INSTRUMENTO ADICIONAL-S[P]"]
            col_SP_CAND  = header_map["CAJA DE CANDADO-S[P]"]

            # Totales (debes calcularlos sin modificar C, D, E)
            col_TOTAL_T = header_map["CAJAS[T]"]     # F
            col_TOTAL_I = header_map["CAJAS-I[T]"]   # M
            col_TOTAL_S = header_map["CAJAS-S[T]"]   # T

            max_row = ws.max_row

            # --- 4) Procesar filas de la plantilla ---
            for r in range(2, max_row + 1):

                sede_pl = limpiar(ws[f"{col_sede}{r}"].value)
                if not sede_pl:
                    continue

                datos = {
                    "INSTRUMENTO": index.get((sede_pl, "INSTRUMENTO"), {"T": 0, "I": 0, "S": 0}),
                    "ADICIONAL":  index.get((sede_pl, "ADICIONAL"),  {"T": 0, "I": 0, "S": 0}),
                    "CANDADO":    index.get((sede_pl, "CANDADO"),    {"T": 0, "I": 0, "S": 0}),
                }

                # =====================================================
                # >>>>> NO MODIFICAR C, D, E  (se dejan igual)
                # =====================================================

                # --- I (colocar valores) ---
                ws[f"{col_I_INSTR}{r}"] = datos["INSTRUMENTO"]["I"]
                ws[f"{col_I_ADIC}{r}"]  = datos["ADICIONAL"]["I"]
                ws[f"{col_I_CAND}{r}"]  = datos["CANDADO"]["I"]

                # --- S (colocar valores) ---
                ws[f"{col_S_INSTR}{r}"] = datos["INSTRUMENTO"]["S"]
                ws[f"{col_S_ADIC}{r}"]  = datos["ADICIONAL"]["S"]
                ws[f"{col_S_CAND}{r}"]  = datos["CANDADO"]["S"]

                # --- Totales (usar columnas C, D, E originales) ---
                ws[f"{col_TOTAL_T}{r}"] = f"=C{r}+D{r}+E{r}"
                ws[f"{col_TOTAL_I}{r}"] = f"={col_I_INSTR}{r}+{col_I_ADIC}{r}+{col_I_CAND}{r}"
                ws[f"{col_TOTAL_S}{r}"] = f"={col_S_INSTR}{r}+{col_S_ADIC}{r}+{col_S_CAND}{r}"

                # --- Porcentajes usando columnas I,T y S ---
                ws[f"{col_IP_INSTR}{r}"] = f"=IF(C{r}=0,1,{col_I_INSTR}{r}/C{r})"
                ws[f"{col_IP_ADIC}{r}"]  = f"=IF(D{r}=0,1,{col_I_ADIC}{r}/D{r})"
                ws[f"{col_IP_CAND}{r}"]  = f"=IF(E{r}=0,1,{col_I_CAND}{r}/E{r})"

                ws[f"{col_SP_INSTR}{r}"] = f"=IF(C{r}=0,1,{col_S_INSTR}{r}/C{r})"
                ws[f"{col_SP_ADIC}{r}"]  = f"=IF(D{r}=0,1,{col_S_ADIC}{r}/D{r})"
                ws[f"{col_SP_CAND}{r}"]  = f"=IF(E{r}=0,1,{col_S_CAND}{r}/E{r})"

            # --- 5) Formato condicional para porcentajes ---
            for nombre, letra in header_map.items():
                if "[P]" in nombre:
                    rango = f"{letra}2:{letra}{max_row}"
                    regla = CellIsRule(
                        operator="lessThan",
                        formula=["1"],
                        font=Font(color="FFFF0000")
                    )
                    ws.conditional_formatting.add(rango, regla)

            # --- 6) Guardar SOLO esta hoja ---
            for hoja in wb.sheetnames:
                if hoja != "CAJAS-SEDE":
                    del wb[hoja]

            out = BytesIO()
            wb.save(out)
            out.seek(0)

            st.session_state["cajas_sede_generada"] = out

        st.success("Hoja CAJAS-SEDE generada correctamente ✔")

    except Exception as e:
        st.error(f"Error al generar CAJAS-SEDE: {e}")

