# funciones_personal.py
# Versión FINAL con detección robusta + formato condicional completo
import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font
from io import BytesIO


# -----------------------------------------------------------
# LIMPIEZA DE TEXTO
# -----------------------------------------------------------
def limpiar(s):
    if pd.isna(s):
        return ""
    return (
        str(s)
        .replace("\xa0", " ")
        .replace("\t", " ")
        .replace("\r", " ")
        .strip()
        .upper()
    )


# -----------------------------------------------------------
# DETECTAR FILA DE ENCABEZADOS
# -----------------------------------------------------------
def detectar_fila_encabezados(df_raw):
    columnas_clave = [
        "SEDE OPERATIVA",
        "LOCAL",
        "CARGO",
        "MÍNIMO REQUERIDO",
        "ASISTENCIA",
    ]

    for i in range(12):
        fila = df_raw.iloc[i].astype(str).apply(limpiar)
        coincidencias = sum(col in fila.values for col in columnas_clave)
        if coincidencias >= 4:
            return i

    raise ValueError("No se pudo detectar la fila de encabezados en ASC-PERSONAL.")


# -----------------------------------------------------------
# CARGAR ASC-PERSONAL
# -----------------------------------------------------------
def _cargar_asc_personal(archivo_asc):
    df_raw = pd.read_excel(archivo_asc, sheet_name="Reporte_Nacional", header=None)
    header_row = detectar_fila_encabezados(df_raw)

    df = pd.read_excel(
        archivo_asc,
        sheet_name="Reporte_Nacional",
        header=header_row
    )

    df.columns = [limpiar(c) for c in df.columns]

    columnas_necesarias = [
        "SEDE OPERATIVA",
        "LOCAL",
        "CARGO",
        "MÍNIMO REQUERIDO",
        "ASISTENCIA",
    ]

    faltan = [c for c in columnas_necesarias if c not in df.columns]
    if faltan:
        raise ValueError(f"Faltan columnas en ASC-PERSONAL: {faltan}")

    df["SEDE OPERATIVA"] = df["SEDE OPERATIVA"].apply(limpiar)
    df["LOCAL"] = df["LOCAL"].apply(limpiar)
    df["CARGO"] = df["CARGO"].apply(limpiar)

    return df


# -----------------------------------------------------------
# MAPEO DE ROLES
# -----------------------------------------------------------
ROLE_MAPPING = {
    "A":   "APLICADOR DE LOCAL DE EVALUACIÓN",
    "AAS": "ASISTENTE ADMINISTRATIVO DE SEDE-INEI",
    "ASL": "AUXILIAR DE SALUD DE LOCAL DE EVALUACIÓN",
    "CAS": "COORDINADOR ADMINISTRATIVO DE SEDE-INEI",
    "CAL": "COORDINADOR ASISTENTE DE LOCAL",
    "CAE": "COORDINADOR DE AULAS DE EVALUACIÓN",
    "CPRS": "COORDINADOR DE PREVENCIÓN DE RIESGOS DE SEGURIDAD DE SEDE",
    "CS":  "COORDINADOR DE SEDE",
    "CLL": "COORDINADOR LÍDER DE LOCAL",
    "CTL": "COORDINADOR TECNOLÓGICO DE LOCAL",
    "CTS": "COORDINADOR TECNOLÓGICO DE SEDE",
    "MS":  "MONITOR DE SEDE",
    "MM":  "MONITOR MINEDU",
    "MR":  "MONITOR REGIONAL",
    "OM":  "OPERADOR DE MANTENIMIENTO DE LOCAL DE EVALUACIÓN",
    "OT":  "OPERADOR TECNOLÓGICO DE LOCAL DE EVALUACIÓN",
    "O":   "ORIENTADOR DE LOCAL DE EVALUACIÓN",
    "PERSONAL DE LIMPIEZA": "PERSONAL DE LIMPIEZA",
    "PERSONAL DE VIGILANCIA DE LA SEDE": "PERSONAL DE VIGILANCIA DE LA SEDE",
    "SLE": "SUPERVISOR DE LOCAL DE EVALUACIÓN",
    "SN":  "SUPERVISOR NACIONAL",
}


# -----------------------------------------------------------
# GENERAR HOJA PERSONAL
# -----------------------------------------------------------
def generar_personal(ruta_plantilla_temp, archivo_asc_personal):

    try:
        with st.spinner("Generando hoja PERSONAL..."):

            df_asc = _cargar_asc_personal(archivo_asc_personal)

            asc_idx = {}
            for _, row in df_asc.iterrows():
                key = (
                    row["SEDE OPERATIVA"],
                    row["LOCAL"],
                    row["CARGO"]
                )
                asc_idx[key] = row

            wb = load_workbook(ruta_plantilla_temp)
            ws = wb["PERSONAL"]

            header_map = {}
            for cell in ws[1]:
                if cell.value:
                    nombre = limpiar(cell.value)
                    if nombre:
                        header_map[nombre] = (cell.column, cell.column_letter)

            if "SEDE" not in header_map or "LOCAL" not in header_map:
                raise ValueError("La plantilla no tiene las columnas SEDE y LOCAL correctamente definidas.")

            col_sede = header_map["SEDE"][1]
            col_local = header_map["LOCAL"][1]

            totals_cols = {}
            base_cols = {}
            perc_cols = {}
            diff_cols = {}

            for name, (idx, let) in header_map.items():

                if name.endswith("[T]"):
                    base = name.replace("[T]", "").strip()
                    totals_cols[base] = name

                elif name.endswith("[P]"):
                    base = name.replace("[P]", "").strip()
                    perc_cols[base] = name

                elif name.endswith("[D]"):
                    base = name.replace("[D]", "").strip()
                    diff_cols[base] = name

                else:
                    if name not in ("N", "SEDE", "LOCAL"):
                        base_cols[name] = name

            max_row = ws.max_row

            # ------------------------------------------------------
            # RELLENAR HOJA PERSONAL
            # ------------------------------------------------------
            for r in range(2, max_row + 1):

                sede = limpiar(ws[f"{col_sede}{r}"].value)
                local = limpiar(ws[f"{col_local}{r}"].value)

                if not sede or not local:
                    continue

                for base, cargo_name in ROLE_MAPPING.items():

                    cargo = limpiar(cargo_name)
                    key = (sede, local, cargo)

                    minimo = 0
                    asistencia = 0

                    if key in asc_idx:
                        minimo = int(asc_idx[key]["MÍNIMO REQUERIDO"] or 0)
                        asistencia = int(asc_idx[key]["ASISTENCIA"] or 0)

                    # TOTAL T
                    if base in totals_cols:
                        colT = header_map[totals_cols[base]][1]
                        ws[f"{colT}{r}"] = minimo

                    # ASISTENCIA
                    if base in base_cols:
                        colA = header_map[base_cols[base]][1]
                        ws[f"{colA}{r}"] = asistencia

                    # PORCENTAJE
                    if base in perc_cols and base in totals_cols and base in base_cols:
                        colp = header_map[perc_cols[base]][1]
                        colT = header_map[totals_cols[base]][1]
                        colA = header_map[base_cols[base]][1]
                        ws[f"{colp}{r}"] = f"=IF({colT}{r}=0,1,{colA}{r}/{colT}{r})"

                    # DIFERENCIA
                    if base in diff_cols and base in totals_cols and base in base_cols:
                        cold = header_map[diff_cols[base]][1]
                        colT = header_map[totals_cols[base]][1]
                        colA = header_map[base_cols[base]][1]
                        ws[f"{cold}{r}"] = f"={colT}{r}-{colA}{r}"

            # ------------------------------------------------------
            # FORMATO CONDICIONAL PORCENTAJE < 100%
            # ------------------------------------------------------
            for base, colname in perc_cols.items():
                col_letter = header_map[colname][1]
                rango = f"{col_letter}2:{col_letter}{max_row}"

                regla_rojo = CellIsRule(
                    operator="lessThan",
                    formula=["1"],
                    stopIfTrue=False,
                    font=Font(color="FFFF0000")
                )
                ws.conditional_formatting.add(rango, regla_rojo)

            # ------------------------------------------------------
            # FORMATO CONDICIONAL DIFERENCIA > 0
            # ------------------------------------------------------
            for base, colname in diff_cols.items():
                col_letter = header_map[colname][1]
                rango = f"{col_letter}2:{col_letter}{max_row}"

                regla_rojo_d = CellIsRule(
                    operator="greaterThan",
                    formula=["0"],
                    stopIfTrue=False,
                    font=Font(color="FFFF0000")
                )
                ws.conditional_formatting.add(rango, regla_rojo_d)

            # ------------------------------------------------------
            # ✔ EXPORTAR SOLO LA HOJA PERSONAL SIN DAÑAR LA PLANTILLA
            # ------------------------------------------------------
            hojas = wb.sheetnames
            for hoja in hojas:
                if hoja != "PERSONAL":
                    del wb[hoja]

            out = BytesIO()
            wb.save(out)
            out.seek(0)

            st.session_state["personal_generada"] = out

        st.success("Hoja PERSONAL generada correctamente ✔")

    except Exception as e:
        st.error(f"Error al generar PERSONAL: {e}")