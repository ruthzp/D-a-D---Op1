import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.properties import CalcProperties
import streamlit as st


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def habilitar_recalculo(wb):
    """Activa recálculo completo al abrir Excel."""
    try:
        wb.calculation_properties.fullCalcOnLoad = True
    except Exception:
        wb.calculation_properties = CalcProperties(fullCalcOnLoad=True)


def cargar_excel_con_encabezado_correcto(file):
    """
    Detecta la fila donde aparece 'Sede Operativa'
    y la usa como fila de encabezado.
    """
    df_raw = pd.read_excel(file, header=None)
    header_row = None

    for i in range(len(df_raw)):
        fila = df_raw.iloc[i].astype(str).str.lower()
        if fila.str.contains("sede operativa").any():
            header_row = i
            break

    if header_row is None:
        raise ValueError("❌ No se encontró la fila de encabezado (Sede Operativa).")

    df = pd.read_excel(file, header=header_row)
    df.columns = df.columns.str.strip()
    return df


# ============================================================
# FUNCIÓN PRINCIPAL — GENERAR OP1
# ============================================================

def generar_op1(base, asc_fa, asc_inst, nom_inst, mindef_inst=None):

    st.info("Procesando hoja OP1...")

    try:
        asc_fa_df = cargar_excel_con_encabezado_correcto(asc_fa)
        asc_inst_df = cargar_excel_con_encabezado_correcto(asc_inst)
        nom_inst_df = cargar_excel_con_encabezado_correcto(nom_inst)

        # MINDEF opcional
        if mindef_inst:
            try:
                mindef_inst_df = cargar_excel_con_encabezado_correcto(mindef_inst)
            except Exception:
                st.warning("⚠ MINDEF - INSTRUMENTOS no válido. Se usará 0.")
                mindef_inst_df = None
        else:
            mindef_inst_df = None

        wb = load_workbook(base)

        if "OP1" not in wb.sheetnames:
            raise ValueError("❌ La plantilla no contiene la hoja OP1.")

        # dejar solo OP1
        for h in wb.sheetnames.copy():
            if h != "OP1":
                del wb[h]

        ws = wb["OP1"]

        actualizar_OP1(ws, asc_fa_df, asc_inst_df, nom_inst_df, mindef_inst_df)

        habilitar_recalculo(wb)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        st.session_state["op1_generada"] = out

        st.success("✅ OP1 generado correctamente.")

    except Exception as e:
        st.error(f"❌ Error al generar OP1: {e}")


# ============================================================
# ACTUALIZAR HOJA OP1
# ============================================================

def actualizar_OP1(ws, asc_fa_df, asc_inst_df, nom_inst_df, mindef_inst_df):

    col_inv = "Inventario en campo"

    def norm(df):
        df.columns = df.columns.str.strip()
        return df

    asc_fa_df = norm(asc_fa_df)
    asc_inst_df = norm(asc_inst_df)
    nom_inst_df = norm(nom_inst_df)
    if mindef_inst_df is not None:
        mindef_inst_df = norm(mindef_inst_df)

    # =====================================================
    # RECORRER FILAS
    # =====================================================
    for r in range(2, ws.max_row + 1):

        sede = str(ws[f"B{r}"].value or "").strip()
        local = str(ws[f"C{r}"].value or "").strip()
        if not sede or not local:
            continue

        # =====================================================
        # ASC — INSTRUMENTOS (O–T)
        # =====================================================

        asc_fil = (
            (asc_inst_df["Sede Operativa"].astype(str).str.strip() == sede)
            & (asc_inst_df["Local"].astype(str).str.strip() == local)
        )

        asc_c = asc_inst_df.loc[
            asc_fil
            & asc_inst_df["Tipo"].astype(str).str.contains(
                "CUADERNILLO DE CONOCIMIENTOS", case=False, na=False
            ),
            col_inv,
        ].sum()

        asc_f = asc_inst_df.loc[
            asc_fil
            & asc_inst_df["Tipo"].astype(str).str.contains(
                "FICHA DE RESPUESTA", case=False, na=False
            ),
            col_inv,
        ].sum()

        ws[f"O{r}"] = asc_c
        ws[f"P{r}"] = asc_f
        ws[f"Q{r}"] = f"=G{r}-O{r}"           # ASC-C[d]
        ws[f"R{r}"] = f"=H{r}-P{r}"           # ASC-F[d]
        ws[f"S{r}"] = f"=IF(G{r}=0,1,O{r}/G{r})"  # ASC-C[p]
        ws[f"T{r}"] = f"=IF(H{r}=0,1,P{r}/H{r})"  # ASC-F[p]

        # =====================================================
        # NOM — INSTRUMENTOS (U–Z)
        # =====================================================

        nom_fil = (
            (nom_inst_df["Sede Operativa"].astype(str).str.strip() == sede)
            & (nom_inst_df["Local"].astype(str).str.strip() == local)
        )

        nom_c = nom_inst_df.loc[
            nom_fil
            & (
                nom_inst_df["Tipo"].astype(str).str.contains(
                    "CUADERNILLO DE HABILIDADES", case=False, na=False
                )
                | nom_inst_df["Tipo"].astype(str).str.contains(
                    "CUADERNILLO DE CONOCIMIENTOS", case=False, na=False
                )
            ),
            col_inv,
        ].sum()

        nom_f = nom_inst_df.loc[
            nom_fil
            & nom_inst_df["Tipo"].astype(str).str.contains(
                "FICHA DE RESPUESTA", case=False, na=False
            ),
            col_inv,
        ].sum()

        ws[f"U{r}"] = nom_c
        ws[f"V{r}"] = nom_f
        ws[f"W{r}"] = f"=I{r}-U{r}"           # NOM-C[d]
        ws[f"X{r}"] = f"=J{r}-V{r}"           # NOM-F[d]
        ws[f"Y{r}"] = f"=IF(I{r}=0,1,U{r}/I{r})"  # NOM-C[p]
        ws[f"Z{r}"] = f"=IF(J{r}=0,1,V{r}/J{r})"  # NOM-F[p]

        # =====================================================
        # MINDEF — INSTRUMENTOS (AA–AF) *opcional*
        # =====================================================

        if mindef_inst_df is not None:
            mindef_fil = (
                (mindef_inst_df["Sede Operativa"].astype(str).str.strip() == sede)
                & (mindef_inst_df["Local"].astype(str).str.strip() == local)
            )

            mindef_c = mindef_inst_df.loc[
                mindef_fil
                & mindef_inst_df["Tipo"].astype(str).str.contains(
                    "CUADERNILLO", case=False, na=False
                ),
                col_inv,
            ].sum()

            mindef_f = mindef_inst_df.loc[
                mindef_fil
                & mindef_inst_df["Tipo"].astype(str).str.contains(
                    "FICHA DE RESPUESTA", case=False, na=False
                ),
                col_inv,
            ].sum()
        else:
            mindef_c = 0
            mindef_f = 0

        ws[f"AA{r}"] = mindef_c
        ws[f"AB{r}"] = mindef_f
        ws[f"AC{r}"] = f"=K{r}-AA{r}"              # MINDEF-C[d]
        ws[f"AD{r}"] = f"=L{r}-AB{r}"              # MINDEF-F[d]
        ws[f"AE{r}"] = f"=IF(K{r}=0,1,AA{r}/K{r})" # MINDEF-C[p]
        ws[f"AF{r}"] = f"=IF(L{r}=0,1,AB{r}/L{r})" # MINDEF-F[p]

        # =====================================================
        # FA — Formatos Auxiliares (AG–AU y AV–BS)
        # =====================================================

        def sumar_fa(tipo_exact):
            return asc_fa_df.loc[
                (asc_fa_df["Sede Operativa"].astype(str).str.strip() == sede)
                & (asc_fa_df["Local"].astype(str).str.strip() == local)
                & (asc_fa_df["Tipo"].astype(str).str.contains(
                    tipo_exact, case=False, na=False
                )),
                col_inv,
            ].sum()

        fa_tipos = {
            "AV": "ACTA DE RECEPCIÓN/DEVOLUCIÓN",
            "AX": "ACTA DE APLICACIÓN DEL AULA",
            "AZ": "LISTA DE ASISTENCIA",
            "BB": "LISTA DE RETIRO DE CUADERNILLOS",
            "BD": "ACTA DE RESPUESTA A OBSERVACIONES DEL DOCENTE",
            "BF": "REGISTRO DE ENTREGA INSTRUMENTOS ADICIONALES",
            "BH": "ACTA DE INCIDENCIAS DEL CAE",
            "BJ": "ACTA DE INCUMPLIMIENTO DE PROCEDIMIENTOS",
            "BL": "ACTA DE INCIDENCIAS DE SALUD",
            "BN": "ACTA DE INCIDENCIAS DEL LOCAL DE EVALUACIÓN",
            "BP": "ACTA FISCAL",
            "BR": "SOBRES",
        }

        for col_letra, texto in fa_tipos.items():
            ws[f"{col_letra}{r}"] = sumar_fa(texto)

        # --------------------------
        # PORCENTAJES / ESTADOS FA
        # --------------------------

        ws[f"AW{r}"] = f"=IF(AJ{r}=0,1,AV{r}/AJ{r})"
        ws[f"AY{r}"] = f"=IF(AK{r}=0,1,AX{r}/AK{r})"
        ws[f"BA{r}"] = f"=IF(AL{r}=0,1,AZ{r}/AL{r})"
        ws[f"BC{r}"] = f"=IF(AM{r}=0,1,BB{r}/AM{r})"
        ws[f"BE{r}"] = f"=IF(AN{r}=0,1,BD{r}/AN{r})"
        ws[f"BG{r}"] = f"=IF(AO{r}=0,1,BF{r}/AO{r})"
        ws[f"BI{r}"] = f"=IF(AP{r}=0,1,BH{r}/AP{r})"

        # ✔ BK y BQ con fórmulas de OK/ERR (no porcentaje)
        ws[f"BK{r}"] = f'=IF(MOD(BJ{r},2)=0,"OK","ERR")'# Acta de incumplimiento de procedimientos
        ws[f"BQ{r}"] = f'=IF(AT{r}=0,0,BP{r}/AT{r})'
        ws[f"BQ{r}"].number_format = "0.00%"  # Acta fiscal

        ws[f"BM{r}"] = f"=IF(AR{r}=0,1,BL{r}/AR{r})"
        ws[f"BO{r}"] = f"=IF(AS{r}=0,1,BN{r}/AS{r})"
        ws[f"BS{r}"] = f"=IF(AU{r}=0,1,BR{r}/AU{r})"

    # =====================================================
    # FORMATO CONDICIONAL PARA [d] (≠ 0 → rojo)
    # =====================================================

    columnas_d = ["Q", "R", "W", "X", "AC", "AD"]

    for col in columnas_d:
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            CellIsRule(
                operator="notEqual",
                formula=["0"],
                font=Font(color="FF0000")
            )
        )

       
    # =====================================================
    # FORMATO CONDICIONAL PARA [p] (< 1 → rojo)
    # (NO incluye BK ni BQ porque son OK/ERR, no proporción)
    # =====================================================

    columnas_p = [
        "S", "T", "Y", "Z", "AE", "AF",
        "AW", "AY", "BA", "BC", "BE", "BG", "BI",
        "BM", "BO", "BS"
    ]

    for col in columnas_p:
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            CellIsRule(
                operator="lessThan",
                formula=["1"],
                font=Font(color="FF0000")
            )
        )

            # =====================================================
    # NUEVO → Formato condicional para BQ < 1 (100%)
    # =====================================================

    ws.conditional_formatting.add(
        f"BQ2:BQ{ws.max_row}",
        CellIsRule(
            operator="lessThan",
            formula=["1"],
            font=Font(color="FF0000")
        )
    )

    # =====================================================
    # FORMATO CONDICIONAL SOLO PARA BK y BQ (ERR → rojo)
    # =====================================================

    err_columns = ["BK", "BQ"]

    for col in err_columns:
        ws.conditional_formatting.add(
            f"{col}2:{col}{ws.max_row}",
            CellIsRule(
                operator="equal",
                formula=['"ERR"'],
                font=Font(color="FF0000")
            )
        )


    return ws