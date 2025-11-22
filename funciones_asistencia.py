import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule


# ---------------------------------------------------------
# Detecta la columna “Sede”
# ---------------------------------------------------------
def detectar_columna_sede(df):
    for c in df.columns:
        if any(x in str(c).lower() for x in [
            "sede", "operativa", "evaluación", "evaluacion",
            "aplicación", "aplicacion"
        ]):
            return c
    raise ValueError("❌ No se encontró una columna de sede válida.")


# ---------------------------------------------------------
# Carga archivos ASC, NOM y MINDEF (antes ACC)
# ---------------------------------------------------------
def cargar_postulantes(file):
    df = pd.read_excel(file, header=None)

    # Buscar fila con cabecera (valor "N" en la primera columna)
    cab = df.index[df.iloc[:, 0].astype(str).str.upper().eq("N")].tolist()
    if not cab:
        raise ValueError("❌ No se encontró la fila de cabecera (valor 'N').")

    df.columns = df.iloc[cab[0]]
    df = df.iloc[cab[0] + 1:].reset_index(drop=True)

    sede_col = detectar_columna_sede(df)
    df = df.rename(columns={sede_col: "Sede"})

    # Convertir campos numéricos
    for c in ["Postulantes", "Asistencia al Local",
              "Asistencia en Aula", "Casos de inconsistencia"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    return df.groupby("Sede", as_index=False).sum()


# ---------------------------------------------------------
# Forzar recálculo en Excel
# ---------------------------------------------------------
def habilitar_recalculo(wb):
    from openpyxl.workbook.properties import CalcProperties
    try:
        wb.calculation_properties.fullCalcOnLoad = True
    except Exception:
        wb.calculation_properties = CalcProperties(fullCalcOnLoad=True)


# ---------------------------------------------------------
# FUNCIÓN PRINCIPAL — GENERAR ASISTENCIA
# ---------------------------------------------------------
def generar_asistencia(base, asc, nom, mindef):
    st.info("Procesando hoja ASISTENCIA...")

    try:
        # === Cargar ASC ===
        asc_df = cargar_postulantes(asc)
        asc_d = asc_df.set_index("Sede").to_dict("index")

        # === Cargar NOM ===
        nom_df = cargar_postulantes(nom)
        nom_d = nom_df.set_index("Sede").to_dict("index")

        # === Cargar MINDEF (opcional) ===
        if mindef:
            mindef_df = cargar_postulantes(mindef)
            mindef_d = mindef_df.set_index("Sede").to_dict("index")
        else:
            mindef_d = {}   # si no hay archivo MINDEF → valores 0

        # === Abrir la plantilla ===
        wb = load_workbook(base)
        if "ASISTENCIA" not in wb.sheetnames:
            raise ValueError("❌ La plantilla no contiene la hoja ASISTENCIA.")

        # Dejar solo la hoja ASISTENCIA
        for h in wb.sheetnames.copy():
            if h != "ASISTENCIA":
                del wb[h]

        ws = wb["ASISTENCIA"]

        # Colores
        rojo = PatternFill("solid", fgColor="FFC7CE")
        verde = PatternFill("solid", fgColor="C6EFCE")

        # ---------------------------------------------------------
        # Llenar cada fila
        # ---------------------------------------------------------
        for r in range(2, ws.max_row + 1):
            sede = str(ws[f"B{r}"].value or "").strip()
            if not sede:
                continue

            get = lambda d, k: d.get(sede, {}).get(k, 0)

            # ----------------------------------
            #  ASC  → columnas E, F, G, H
            # ----------------------------------
            ws[f"E{r}"].value = get(asc_d, "Postulantes")
            ws[f"F{r}"].value = get(asc_d, "Asistencia al Local")
            ws[f"G{r}"].value = get(asc_d, "Asistencia en Aula")
            ws[f"H{r}"].value = get(asc_d, "Casos de inconsistencia")

            # ----------------------------------
            #  NOM → columnas I, J, K, L
            # ----------------------------------
            ws[f"I{r}"].value = get(nom_d, "Postulantes")
            ws[f"J{r}"].value = get(nom_d, "Asistencia al Local")
            ws[f"K{r}"].value = get(nom_d, "Asistencia en Aula")
            ws[f"L{r}"].value = get(nom_d, "Casos de inconsistencia")

            # ----------------------------------
            #  MINDEF (opcional) → M, N, O, P
            # ----------------------------------
            ws[f"M{r}"].value = get(mindef_d, "Postulantes")
            ws[f"N{r}"].value = get(mindef_d, "Asistencia al Local")
            ws[f"O{r}"].value = get(mindef_d, "Asistencia en Aula")
            ws[f"P{r}"].value = get(mindef_d, "Casos de inconsistencia")

            # ----------------------------------
            #  TOTALES
            # ----------------------------------

            # Q = Total Postulantes (ASC + NOM + MINDEF)
            ws[f"Q{r}"].value = f"=E{r}+I{r}+M{r}"

            # R = Total Local
            ws[f"R{r}"].value = f"=F{r}+J{r}+N{r}"

            # S = Total Aula
            ws[f"S{r}"].value = f"=G{r}+K{r}+O{r}"

            # T = Total Inconsistencias
            ws[f"T{r}"].value = f"=H{r}+L{r}+P{r}"

            # ----------------------------------
            #  ESTADO (U)
            # ----------------------------------
            ws[f"U{r}"].value = f'=IF($D{r}=$T{r},"OK","ERR")'

        # ---------------------------------------------------------
        # FORMATO CONDICIONAL (U)
        # ---------------------------------------------------------
        ws.conditional_formatting.add(
            f"U2:U{ws.max_row}",
            CellIsRule("equal", ['"ERR"'], fill=rojo)
        )
        ws.conditional_formatting.add(
            f"U2:U{ws.max_row}",
            CellIsRule("equal", ['"OK"'], fill=verde)
        )

        # Recalcular al abrir en Excel
        habilitar_recalculo(wb)

        # Salida
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        st.session_state["asistencia_generada"] = out

        st.success("✅ Hoja ASISTENCIA generada correctamente.")

    except Exception as e:
        st.error(f"❌ Error al generar ASISTENCIA: {e}")